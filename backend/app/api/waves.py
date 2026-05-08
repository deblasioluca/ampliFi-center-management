"""Wave API endpoints (section 11.7)."""

from __future__ import annotations

import logging
import secrets
import threading
from datetime import UTC, datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.api.deps import PaginationParams, get_current_user, pagination, require_role
from app.infra.db.session import get_db
from app.models.core import (
    AnalysisRun,
    AppUser,
    CenterMapping,
    CenterProposal,
    Entity,
    HierarchyLeaf,
    LegacyCostCenter,
    ReviewItem,
    ReviewScope,
    TargetCostCenter,
    TargetProfitCenter,
    Wave,
    WaveEntity,
    WaveHierarchyScope,
    WaveTemplate,
)

log = logging.getLogger(__name__)


def _resolve_entity_by_ccode(db: Session, ccode: str) -> Entity | None:
    """Resolve an Entity by its ccode, picking the right one when multiple exist.

    The ``Entity`` table has a unique constraint on ``(scope, ccode)`` —
    so the same ccode can legitimately appear for both the legacy SAP
    source ("cleanup" scope) and target/explorer scopes. The wave UI
    selects entities by ccode alone (the picker doesn't expose scope),
    which under ``.scalar_one_or_none()`` would either return None or
    raise ``MultipleResultsFound`` depending on how many rows exist —
    both failure modes silently dropping the entity on wave creation.

    This helper picks deterministically: prefer the cleanup-scope row
    (which is the analysis source), then fall back to whatever exists
    so a wave can still be created in test environments where only
    target rows are loaded.
    """
    rows = (
        db.execute(select(Entity).where(Entity.ccode == ccode).order_by(Entity.scope))
        .scalars()
        .all()
    )
    if not rows:
        return None
    # Prefer cleanup scope (default scope for legacy SAP source), then
    # any other row. This matches "the wave's entities are the ones
    # whose CCs we'll cleanse" — the analysis pipeline reads from the
    # cleanup-scope CCs.
    for r in rows:
        if r.scope == "cleanup":
            return r
    return rows[0]


router = APIRouter()

logger = __import__("structlog").get_logger()


def _send_scope_invitation(
    scope: ReviewScope, wave: Wave, db: Session, *, is_reminder: bool = False
) -> None:
    """Best-effort email invitation to reviewer.

    When ``is_reminder=True``, sends the ``review_reminder`` template (with
    progress counters) instead of the initial ``review_invitation``. The
    reminder template explicitly tells reviewers how many items they've
    decided already, so they don't read it as a duplicate of the first
    invitation.
    """
    from app.models.core import AppConfig

    try:
        cfg = db.execute(select(AppConfig).where(AppConfig.key == "email")).scalar_one_or_none()
        if not cfg or not cfg.value:
            return
        email_cfg = cfg.value
        from app.infra.email.engine import EmailEngine

        engine = EmailEngine(
            host=email_cfg.get("host", "localhost"),
            port=int(email_cfg.get("port", 1025)),
            username=email_cfg.get("username", ""),
            password=email_cfg.get("password", ""),
            use_tls=email_cfg.get("tls", "none") != "none",
            from_address=email_cfg.get("from_address", "noreply@amplifi.dev"),
        )
        item_count = (
            db.execute(
                select(func.count(ReviewItem.id)).where(ReviewItem.scope_id == scope.id)
            ).scalar()
            or 0
        )

        if is_reminder:
            decided_count = (
                db.execute(
                    select(func.count(ReviewItem.id)).where(
                        ReviewItem.scope_id == scope.id,
                        ReviewItem.decision != "PENDING",
                    )
                ).scalar()
                or 0
            )
            engine.send(
                to=scope.reviewer_email,
                template_name="review_reminder",
                context={
                    "reviewer_name": scope.reviewer_name or scope.reviewer_email,
                    "wave_name": wave.name,
                    "review_url": f"/review/{scope.token}",
                    "reviewed_count": decided_count,
                    "total_count": item_count,
                    "deadline": (
                        str(scope.token_expires_at)[:10] if scope.token_expires_at else "soon"
                    ),
                },
            )
            logger.info("email.reminder_sent", scope_id=scope.id, to=scope.reviewer_email)
        else:
            engine.send(
                to=scope.reviewer_email,
                template_name="review_invitation",
                context={
                    "reviewer_name": scope.reviewer_name or scope.reviewer_email,
                    "wave_name": wave.name,
                    "review_url": f"/review/{scope.token}",
                    "expires_at": (
                        str(scope.token_expires_at)[:10] if scope.token_expires_at else "N/A"
                    ),
                    "scope_name": scope.name,
                    "item_count": item_count,
                },
            )
            logger.info("email.invitation_sent", scope_id=scope.id, to=scope.reviewer_email)
    except Exception:
        logger.warning("email.invitation_failed", scope_id=scope.id, exc_info=True)


VALID_TRANSITIONS = {
    "draft": ["analysing", "cancelled"],
    "analysing": ["proposed", "draft", "cancelled"],
    "proposed": ["locked", "draft", "cancelled"],
    "locked": ["in_review", "proposed"],
    "in_review": ["signed_off"],
    "signed_off": ["closed"],
    "closed": [],
    "cancelled": [],
}


class HierarchyScopeIn(BaseModel):
    hierarchy_id: int
    node_setname: str


class WaveCreate(BaseModel):
    code: str
    name: str
    description: str | None = None
    is_full_scope: bool = False
    exclude_prior: bool = True
    entity_ccodes: list[str] = []
    hierarchy_scopes: list[HierarchyScopeIn] = []


class WaveUpdate(BaseModel):
    name: str | None = None
    description: str | None = None


class WaveOut(BaseModel):
    id: int
    code: str
    name: str
    description: str | None
    status: str
    is_full_scope: bool
    exclude_prior: bool
    is_archived: bool = False
    entity_count: int = 0
    config: dict | None = None

    model_config = {"from_attributes": True}


@router.get("")
def list_waves(
    db: Session = Depends(get_db),
    user: AppUser = Depends(get_current_user),
    pag: PaginationParams = Depends(pagination),
    status: str | None = None,
    archived: str = "false",
) -> dict:
    """List waves, optionally filtered by status and archive flag.

    The ``archived`` parameter selects which slice of the universe the
    caller wants to see:
    * ``"false"`` (default): only non-archived waves. This is what the
      cockpit's wave list and the analytics dashboard see — the
      everyday active set.
    * ``"true"``: only archived waves. Used by the admin archive page
      so the operator can review what's been put away (and delete it
      from there).
    * ``"all"``: both archived and non-archived. Useful for system-
      wide audits / debugging.

    Anything other than these three values is treated as ``"false"``
    so a typo doesn't accidentally surface archived rows.
    """
    query = select(Wave).order_by(Wave.created_at.desc())
    archived_mode = archived if archived in ("false", "true", "all") else "false"
    if archived_mode == "false":
        query = query.where(Wave.is_archived.is_(False))
    elif archived_mode == "true":
        query = query.where(Wave.is_archived.is_(True))
    if status:
        status_list = [s.strip() for s in status.split(",") if s.strip()]
        if len(status_list) == 1:
            query = query.where(Wave.status == status_list[0])
        else:
            query = query.where(Wave.status.in_(status_list))
    total_q = select(func.count(Wave.id))
    if archived_mode == "false":
        total_q = total_q.where(Wave.is_archived.is_(False))
    elif archived_mode == "true":
        total_q = total_q.where(Wave.is_archived.is_(True))
    if status:
        status_list = [s.strip() for s in status.split(",") if s.strip()]
        if len(status_list) == 1:
            total_q = total_q.where(Wave.status == status_list[0])
        else:
            total_q = total_q.where(Wave.status.in_(status_list))
    total = db.execute(total_q).scalar() or 0
    waves = db.execute(query.offset((pag.page - 1) * pag.size).limit(pag.size)).scalars().all()
    items = []
    for w in waves:
        ec = (
            db.execute(select(func.count(WaveEntity.id)).where(WaveEntity.wave_id == w.id)).scalar()
            or 0
        )
        out = WaveOut(
            id=w.id,
            code=w.code,
            name=w.name,
            description=w.description,
            status=w.status,
            is_full_scope=w.is_full_scope,
            exclude_prior=w.exclude_prior,
            is_archived=w.is_archived,
            entity_count=ec,
            config=w.config,
        )
        items.append(out.model_dump())
    return {"total": total, "page": pag.page, "size": pag.size, "items": items}


@router.post("")
def create_wave(
    body: WaveCreate,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> WaveOut:
    existing = db.execute(select(Wave).where(Wave.code == body.code)).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail="Wave code already exists")
    wave = Wave(
        code=body.code,
        name=body.name,
        description=body.description,
        is_full_scope=body.is_full_scope,
        exclude_prior=body.exclude_prior,
        created_by=user.id,
    )
    db.add(wave)
    db.flush()
    # Track which ccodes we couldn't resolve so the response can warn
    # the user — silent loss was the bug behind "I selected entities
    # but the wave has 0 entities" reports.
    missing_ccodes: list[str] = []
    for ccode in body.entity_ccodes:
        entity = _resolve_entity_by_ccode(db, ccode)
        if entity:
            db.add(WaveEntity(wave_id=wave.id, entity_id=entity.id))
        else:
            missing_ccodes.append(ccode)
            log.warning("create_wave: ccode '%s' not found in Entity table", ccode)
    for hs in body.hierarchy_scopes:
        db.add(
            WaveHierarchyScope(
                wave_id=wave.id,
                hierarchy_id=hs.hierarchy_id,
                node_setname=hs.node_setname,
            )
        )
    db.commit()
    db.refresh(wave)
    ec = (
        db.execute(select(func.count(WaveEntity.id)).where(WaveEntity.wave_id == wave.id)).scalar()
        or 0
    )
    out = WaveOut(
        id=wave.id,
        code=wave.code,
        name=wave.name,
        description=wave.description,
        status=wave.status,
        is_full_scope=wave.is_full_scope,
        exclude_prior=wave.exclude_prior,
        is_archived=wave.is_archived,
        entity_count=ec,
        config=wave.config,
    )
    # Inject the diagnostic into the response. Pydantic strips unknown
    # fields by default, so we serialise + augment manually.
    payload = out.model_dump()
    if missing_ccodes:
        payload["missing_ccodes"] = missing_ccodes
    return payload


# --- Wave Templates (must be defined before /{wave_id} to avoid shadowing) ---


class TemplateCreate(BaseModel):
    name: str
    description: str | None = None
    config: dict | None = None
    is_full_scope: bool = False
    exclude_prior: bool = True
    entity_ccodes: list[str] | None = None


class TemplateOut(BaseModel):
    id: int
    name: str
    description: str | None
    config: dict | None
    is_full_scope: bool
    exclude_prior: bool
    entity_ccodes: list | None

    model_config = {"from_attributes": True}


@router.get("/templates")
def list_templates(db: Session = Depends(get_db)) -> list[TemplateOut]:
    rows = db.execute(select(WaveTemplate).order_by(WaveTemplate.name)).scalars().all()
    return [TemplateOut.model_validate(r) for r in rows]


@router.post("/templates")
def create_template(
    body: TemplateCreate,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin")),
) -> TemplateOut:
    t = WaveTemplate(
        name=body.name,
        description=body.description,
        config=body.config,
        is_full_scope=body.is_full_scope,
        exclude_prior=body.exclude_prior,
        entity_ccodes=body.entity_ccodes,
        created_by=user.id,
    )
    db.add(t)
    db.commit()
    db.refresh(t)
    return TemplateOut.model_validate(t)


@router.delete("/templates/{template_id}")
def delete_template(
    template_id: int,
    db: Session = Depends(get_db),
    _user: AppUser = Depends(require_role("admin")),
) -> dict:
    t = db.get(WaveTemplate, template_id)
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    db.delete(t)
    db.commit()
    return {"deleted": True}


@router.post("/templates/{template_id}/create-wave")
def create_wave_from_template(
    template_id: int,
    code: str,
    name: str,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    """Create a new wave from an existing template."""
    tpl = db.get(WaveTemplate, template_id)
    if not tpl:
        raise HTTPException(status_code=404, detail="Template not found")
    wave = Wave(
        code=code,
        name=name,
        description=tpl.description,
        status="draft",
        is_full_scope=tpl.is_full_scope,
        exclude_prior=tpl.exclude_prior,
        config=tpl.config,
        created_by=user.id,
    )
    db.add(wave)
    db.flush()
    if tpl.entity_ccodes:
        entities = (
            db.execute(select(Entity).where(Entity.ccode.in_(tpl.entity_ccodes))).scalars().all()
        )
        for ent in entities:
            db.add(WaveEntity(wave_id=wave.id, entity_id=ent.id))
    db.commit()
    db.refresh(wave)
    return {"id": wave.id, "code": wave.code, "status": wave.status}


@router.get("/review-scopes")
def my_review_scopes(
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager", "reviewer")),
) -> list[dict]:
    """List review scopes assigned to the current user."""
    conditions = [ReviewScope.reviewer_user_id == user.id]
    if user.email:
        conditions.append(ReviewScope.reviewer_email == user.email)
    if user.display_name:
        conditions.append(ReviewScope.reviewer_name == user.display_name)
    if user.username:
        conditions.append(ReviewScope.reviewer_name == user.username)
    scopes = (
        db.execute(select(ReviewScope).where(or_(*conditions)).order_by(ReviewScope.id.desc()))
        .scalars()
        .all()
    )
    result = []
    for s in scopes:
        total = (
            db.execute(
                select(func.count(ReviewItem.id)).where(ReviewItem.scope_id == s.id)
            ).scalar()
            or 0
        )
        decided = (
            db.execute(
                select(func.count(ReviewItem.id)).where(
                    ReviewItem.scope_id == s.id,
                    ReviewItem.decision != "PENDING",
                )
            ).scalar()
            or 0
        )
        wave = s.wave
        result.append(
            {
                "id": s.id,
                "name": s.name,
                "scope_type": s.scope_type,
                "status": s.status,
                "token": s.token,
                "total_items": total,
                "decided_items": decided,
                "wave_name": wave.name if wave else None,
            }
        )
    return result


@router.get("/{wave_id}")
def get_wave(wave_id: int, db: Session = Depends(get_db)) -> WaveOut:
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    ec = (
        db.execute(select(func.count(WaveEntity.id)).where(WaveEntity.wave_id == wave.id)).scalar()
        or 0
    )
    return WaveOut(
        id=wave.id,
        code=wave.code,
        name=wave.name,
        description=wave.description,
        status=wave.status,
        is_full_scope=wave.is_full_scope,
        exclude_prior=wave.exclude_prior,
        entity_count=ec,
        config=wave.config,
    )


@router.patch("/{wave_id}")
def update_wave(
    wave_id: int,
    body: WaveUpdate,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> WaveOut:
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if body.name is not None:
        wave.name = body.name
    if body.description is not None:
        wave.description = body.description
    db.commit()
    db.refresh(wave)
    ec = (
        db.execute(select(func.count(WaveEntity.id)).where(WaveEntity.wave_id == wave.id)).scalar()
        or 0
    )
    return WaveOut(
        id=wave.id,
        code=wave.code,
        name=wave.name,
        description=wave.description,
        status=wave.status,
        is_full_scope=wave.is_full_scope,
        exclude_prior=wave.exclude_prior,
        entity_count=ec,
        config=wave.config,
    )


@router.post("/{wave_id}/cancel")
def cancel_wave(
    wave_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if "cancelled" not in VALID_TRANSITIONS.get(wave.status, []):
        raise HTTPException(status_code=409, detail=f"Cannot cancel wave in status {wave.status}")
    old_status = wave.status
    wave.status = "cancelled"
    from app.domain.audit import write_audit

    write_audit(
        db,
        action="wave.cancel",
        entity_type="wave",
        entity_id=wave.id,
        actor_id=user.id,
        actor_email=user.email or user.username,
        before={"status": old_status},
        after={"status": "cancelled"},
    )
    db.commit()
    return {"status": "cancelled"}


@router.post("/{wave_id}/proposal/lock")
def lock_proposal(
    wave_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.status != "proposed":
        raise HTTPException(status_code=409, detail="Wave must be in proposed status to lock")
    wave.status = "locked"
    wave.locked_at = datetime.now(UTC)
    from app.domain.audit import write_audit

    write_audit(
        db,
        action="wave.lock",
        entity_type="wave",
        entity_id=wave.id,
        actor_id=user.id,
        actor_email=user.email or user.username,
        before={"status": "proposed"},
        after={"status": "locked"},
    )
    db.commit()
    return {"status": "locked"}


@router.post("/{wave_id}/proposal/unlock")
def unlock_proposal(
    wave_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.status != "locked":
        raise HTTPException(status_code=409, detail="Wave must be locked to unlock")
    scope_count = (
        db.execute(
            select(func.count(ReviewScope.id)).where(
                ReviewScope.wave_id == wave.id,
                ReviewScope.status == "invited",
            )
        ).scalar()
        or 0
    )
    if scope_count > 0:
        raise HTTPException(status_code=409, detail="Cannot unlock: review scopes already invited")
    wave.status = "proposed"
    wave.locked_at = None
    db.commit()
    return {"status": "proposed"}


@router.post("/{wave_id}/signoff")
def signoff_wave(
    wave_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.status != "in_review":
        raise HTTPException(status_code=409, detail="Wave must be in_review to sign off")

    # Completeness check: block if any review items are still PENDING
    pending_count = (
        db.execute(
            select(func.count(ReviewItem.id)).where(
                ReviewItem.scope_id.in_(
                    select(ReviewScope.id).where(ReviewScope.wave_id == wave.id)
                ),
                ReviewItem.decision == "PENDING",
            )
        ).scalar()
        or 0
    )
    if pending_count > 0:
        raise HTTPException(
            status_code=409,
            detail=f"Cannot sign off: {pending_count} review items still PENDING",
        )

    wave.status = "signed_off"
    wave.signed_off_at = datetime.now(UTC)
    from app.domain.audit import write_audit

    write_audit(
        db,
        action="wave.signoff",
        entity_type="wave",
        entity_id=wave.id,
        actor_id=user.id,
        actor_email=user.email or user.username,
        before={"status": "in_review"},
        after={"status": "signed_off"},
    )
    db.commit()
    return {"status": "signed_off"}


@router.post("/{wave_id}/close")
def close_wave(
    wave_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.status != "signed_off":
        raise HTTPException(status_code=409, detail="Wave must be signed_off to close")

    wave.status = "closed"
    wave.closed_at = datetime.now(UTC)
    db.commit()
    return {"status": "closed"}


@router.get("/{wave_id}/entities")
def list_wave_entities(
    wave_id: int,
    db: Session = Depends(get_db),
) -> dict:
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    we_rows = db.execute(select(WaveEntity).where(WaveEntity.wave_id == wave_id)).scalars().all()
    items = []
    for we in we_rows:
        entity = db.get(Entity, we.entity_id)
        if entity:
            cc_count = (
                db.execute(
                    select(func.count(LegacyCostCenter.id)).where(
                        LegacyCostCenter.ccode == entity.ccode
                    )
                ).scalar()
                or 0
            )
            items.append(
                {
                    "entity_id": entity.id,
                    "ccode": entity.ccode,
                    "name": entity.name,
                    "region": entity.region,
                    "cost_centers": cc_count,
                }
            )
    return {"wave_id": wave_id, "items": items}


@router.post("/{wave_id}/entities")
def add_wave_entities(
    wave_id: int,
    body: dict,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.status not in ("draft", "analysing"):
        raise HTTPException(status_code=409, detail="Cannot modify entities in this state")
    ccodes = body.get("ccodes", [])
    added = 0
    missing_ccodes: list[str] = []
    for ccode in ccodes:
        entity = _resolve_entity_by_ccode(db, ccode)
        if not entity:
            missing_ccodes.append(ccode)
            log.warning("add_wave_entities: ccode '%s' not found", ccode)
            continue
        existing = db.execute(
            select(WaveEntity).where(
                WaveEntity.wave_id == wave_id,
                WaveEntity.entity_id == entity.id,
            )
        ).scalar_one_or_none()
        if not existing:
            db.add(WaveEntity(wave_id=wave_id, entity_id=entity.id))
            added += 1
    db.commit()
    return {"added": added, "missing_ccodes": missing_ccodes}


@router.get("/{wave_id}/hierarchy-scopes")
def list_wave_hierarchy_scopes(
    wave_id: int,
    db: Session = Depends(get_db),
) -> dict:
    """List the hierarchy nodes that bound a wave's scope.

    A wave's scope can be defined by a list of entities (company codes,
    via WaveEntity rows) AND/OR a list of hierarchy nodes — pick all
    cost centers that fall under any of those nodes. The hierarchy
    side was previously written by ``create_wave`` but never read back
    by any endpoint, so the wave-detail UI couldn't display existing
    scopes or remove them. This GET surfaces them.
    """
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    rows = (
        db.execute(select(WaveHierarchyScope).where(WaveHierarchyScope.wave_id == wave_id))
        .scalars()
        .all()
    )
    items = [
        {
            "id": r.id,
            "wave_id": r.wave_id,
            "hierarchy_id": r.hierarchy_id,
            "node_setname": r.node_setname,
        }
        for r in rows
    ]
    return {"wave_id": wave_id, "items": items}


class WaveHierarchyScopeIn(BaseModel):
    hierarchy_id: int
    node_setname: str


@router.post("/{wave_id}/hierarchy-scopes")
def add_wave_hierarchy_scope(
    wave_id: int,
    body: WaveHierarchyScopeIn,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    """Add a hierarchy node to the wave's scope.

    Replaces the previous (broken) approach of ``PATCH /api/waves/{id}``
    with a ``config: {hierarchy_scope: [...]}`` body — Pydantic was
    silently dropping the unknown ``config`` field on WaveUpdate, so
    nothing was ever persisted. This endpoint writes to the
    ``WaveHierarchyScope`` table directly, which the analysis pipeline
    now reads when filtering cost centers by scope.

    Idempotent on (hierarchy_id, node_setname): re-adding the same
    node is a no-op and returns the existing row.
    """
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.status not in ("draft", "analysing"):
        raise HTTPException(status_code=409, detail="Cannot modify scope in this state")
    existing = db.execute(
        select(WaveHierarchyScope).where(
            WaveHierarchyScope.wave_id == wave_id,
            WaveHierarchyScope.hierarchy_id == body.hierarchy_id,
            WaveHierarchyScope.node_setname == body.node_setname,
        )
    ).scalar_one_or_none()
    if existing:
        return {
            "id": existing.id,
            "wave_id": existing.wave_id,
            "hierarchy_id": existing.hierarchy_id,
            "node_setname": existing.node_setname,
            "created": False,
        }
    row = WaveHierarchyScope(
        wave_id=wave_id,
        hierarchy_id=body.hierarchy_id,
        node_setname=body.node_setname,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {
        "id": row.id,
        "wave_id": row.wave_id,
        "hierarchy_id": row.hierarchy_id,
        "node_setname": row.node_setname,
        "created": True,
    }


@router.delete("/{wave_id}/hierarchy-scopes/{scope_id}")
def delete_wave_hierarchy_scope(
    wave_id: int,
    scope_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    """Remove a hierarchy node from the wave's scope.

    Used when the operator changes their mind about which subtree
    bounds the wave. The scope is gone from the next analysis run; any
    in-flight runs aren't affected.
    """
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.status not in ("draft", "analysing"):
        raise HTTPException(status_code=409, detail="Cannot modify scope in this state")
    row = db.get(WaveHierarchyScope, scope_id)
    if not row or row.wave_id != wave_id:
        raise HTTPException(status_code=404, detail="Scope not found on this wave")
    db.delete(row)
    db.commit()
    return {"deleted": True, "id": scope_id}


@router.get("/{wave_id}/runs")
def list_wave_runs(
    wave_id: int,
    db: Session = Depends(get_db),
) -> dict:
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    runs = (
        db.execute(
            select(AnalysisRun)
            .where(AnalysisRun.wave_id == wave_id)
            .order_by(AnalysisRun.created_at.desc())
        )
        .scalars()
        .all()
    )
    return {
        "wave_id": wave_id,
        "items": [
            {
                "id": r.id,
                "config_id": r.config_id,
                "status": r.status,
                "kpis": r.kpis,
                "started_at": str(r.started_at) if r.started_at else None,
                "finished_at": str(r.finished_at) if r.finished_at else None,
            }
            for r in runs
        ],
    }


class V1AnalysisParams(BaseModel):
    config_id: int | None = None
    mode: str = "simulation"
    label: str | None = None
    excluded_scopes: list[int] | None = None


@router.post("/{wave_id}/analyse")
def run_analysis(
    wave_id: int,
    params: V1AnalysisParams | None = None,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    """Execute decision tree analysis on wave's cost centers."""
    from app.services.analysis import execute_analysis, get_or_create_default_config

    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.status not in ("draft", "analysing", "proposed"):
        raise HTTPException(
            status_code=409,
            detail=f"Cannot analyse wave in status {wave.status}",
        )

    config_id = params.config_id if params else None
    if config_id is None:
        config = get_or_create_default_config(db)
        config_id = config.id

    sim_mode = params.mode if params else "simulation"
    sim_label = params.label if params else None
    excl = [str(x) for x in (params.excluded_scopes or [])] if params else None

    try:
        run = execute_analysis(
            wave_id,
            config_id,
            user.id,
            db,
            mode=sim_mode,
            label=sim_label,
            excluded_scopes=excl,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Analysis failed: {e}") from None

    return {
        "run_id": run.id,
        "status": run.status,
        "mode": run.mode,
        "label": run.label,
        "kpis": run.kpis,
        "started_at": str(run.started_at) if run.started_at else None,
        "finished_at": str(run.finished_at) if run.finished_at else None,
    }


@router.post("/{wave_id}/propose/{run_id}")
def set_preferred_run(
    wave_id: int,
    run_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    run = db.get(AnalysisRun, run_id)
    if not run or run.wave_id != wave_id:
        raise HTTPException(status_code=404, detail="Run not found in this wave")
    if run.status != "completed":
        raise HTTPException(status_code=409, detail="Run must be completed")
    if wave.status not in ("analysing", "proposed"):
        raise HTTPException(
            status_code=409,
            detail=f"Cannot propose from status {wave.status}",
        )
    wave.status = "proposed"
    wave.preferred_run_id = run_id
    wave.config = {**(wave.config or {}), "preferred_run_id": run_id}
    db.commit()
    return {"status": "proposed", "preferred_run_id": run_id}


@router.post("/{wave_id}/reset-proposals")
def reset_proposals(
    wave_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    """Delete all proposals for a wave and release allocated IDs.

    Used when re-running analysis — clears old proposals and recycles
    any CC/PC IDs allocated from the NamingPool back to the pool.
    """
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.status in ("locked", "in_review", "signed_off", "closed", "cancelled"):
        raise HTTPException(
            status_code=409,
            detail=f"Cannot reset proposals: wave is {wave.status}",
        )

    from app.domain.proposal.service import release_proposal_ids

    proposals = (
        db.execute(select(CenterProposal).join(AnalysisRun).where(AnalysisRun.wave_id == wave_id))
        .scalars()
        .all()
    )

    released_ids = 0
    for p in proposals:
        released_ids += release_proposal_ids(p.id, db)
        db.delete(p)

    if wave.status == "proposed":
        wave.status = "analysing"

    from app.domain.audit import write_audit

    write_audit(
        db,
        action="wave.proposals_reset",
        entity_type="wave",
        entity_id=wave.id,
        actor_id=user.id,
        actor_email=user.email or user.username,
        after={
            "proposals_deleted": len(proposals),
            "ids_released": released_ids,
        },
    )
    db.commit()
    return {
        "proposals_deleted": len(proposals),
        "ids_released": released_ids,
    }


@router.delete("/{wave_id}")
def delete_wave(
    wave_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin")),
) -> dict:
    """Delete a wave and all associated data (proposals, scopes, runs).

    Two paths to deletion:
    * **Active waves** (status in {draft, analysing, proposed, locked,
      in_review}): admin can delete to abandon work-in-progress. The
      previous in_review block has been lifted — operators have asked
      to abort waves stuck in review, and the audit log captures the
      delete event so the action remains traceable.
    * **Terminal waves** (status in {signed_off, closed, cancelled}):
      cannot be deleted directly — the operator must first archive the
      wave (POST /api/waves/{id}/archive). Archive-then-delete is the
      two-step pattern that protects audit history: completed waves
      stay visible in the admin archive view until someone explicitly
      decides they're not needed anymore.
    """
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")

    # Terminal waves require archive-first as a deliberate friction
    # against accidental deletion of completed work.
    if wave.status in ("signed_off", "closed", "cancelled") and not wave.is_archived:
        msg = (
            f"Wave is in terminal status '{wave.status}'. "
            "Archive it first via POST /api/waves/{id}/archive, "
            "then remove it via the admin archive view."
        )
        raise HTTPException(status_code=409, detail=msg)

    from app.domain.proposal.service import release_proposal_ids

    proposals = (
        db.execute(select(CenterProposal).join(AnalysisRun).where(AnalysisRun.wave_id == wave_id))
        .scalars()
        .all()
    )
    for p in proposals:
        release_proposal_ids(p.id, db)

    wave_code = wave.code
    db.delete(wave)
    from app.domain.audit import write_audit

    write_audit(
        db,
        action="wave.deleted",
        entity_type="wave",
        entity_id=wave_id,
        actor_id=user.id,
        actor_email=user.email or user.username,
        after={"wave_code": wave_code, "was_archived": wave.is_archived},
    )
    db.commit()
    return {"status": "deleted"}


@router.post("/{wave_id}/archive")
def archive_wave(
    wave_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin")),
) -> dict:
    """Archive a terminal wave so it disappears from the active list.

    Only allowed on waves in terminal status (``signed_off``, ``closed``,
    ``cancelled``). Archiving an active wave doesn't make sense — the
    point of archive is "I'm done with this wave, declutter my view";
    if the wave isn't done, the right action is to either continue it
    or delete it outright.

    Once archived the wave is hidden from ``GET /api/waves`` (and from
    the cockpit's wave list, the analytics dashboard scope picker, etc.)
    Re-surface it via the admin archive view, which queries
    ``GET /api/waves?archived=true`` to show the archived slice.
    """
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.is_archived:
        raise HTTPException(status_code=409, detail="Wave is already archived")
    if wave.status not in ("signed_off", "closed", "cancelled"):
        raise HTTPException(
            status_code=409,
            detail=(
                f"Cannot archive wave in status '{wave.status}'. "
                "Only terminal waves (signed_off, closed, cancelled) can be archived. "
                "If you want to abandon a non-terminal wave, delete it instead."
            ),
        )
    wave.is_archived = True

    from app.domain.audit import write_audit

    write_audit(
        db,
        action="wave.archived",
        entity_type="wave",
        entity_id=wave_id,
        actor_id=user.id,
        actor_email=user.email or user.username,
        after={"wave_code": wave.code, "status": wave.status},
    )
    db.commit()
    return {"status": "archived", "wave_id": wave_id}


@router.post("/{wave_id}/unarchive")
def unarchive_wave(
    wave_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin")),
) -> dict:
    """Move a wave back from the archive into the active list.

    Useful when a completed wave needs to be referenced or re-opened
    for audit/reporting work after it was put away. Only changes the
    archive flag; the wave's status stays whatever it was when it was
    archived.
    """
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if not wave.is_archived:
        raise HTTPException(status_code=409, detail="Wave is not archived")
    wave.is_archived = False

    from app.domain.audit import write_audit

    write_audit(
        db,
        action="wave.unarchived",
        entity_type="wave",
        entity_id=wave_id,
        actor_id=user.id,
        actor_email=user.email or user.username,
        after={"wave_code": wave.code, "status": wave.status},
    )
    db.commit()
    return {"status": "unarchived", "wave_id": wave_id}


class ReviewScopeCreate(BaseModel):
    name: str
    scope_type: str = "entity"
    scope_filter: dict = {}
    reviewer_name: str | None = None
    reviewer_email: str | None = None
    reviewer_user_id: int | None = None


@router.post("/{wave_id}/scopes")
def create_review_scope(
    wave_id: int,
    body: ReviewScopeCreate,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.status not in ("proposed", "locked", "in_review"):
        raise HTTPException(
            status_code=409,
            detail="Wave must be proposed, locked, or in_review to create scopes",
        )

    token = secrets.token_urlsafe(32)
    expires = datetime.now(UTC) + timedelta(days=30)

    # Resolve reviewer_user_id: prefer explicit ID, then lookup by email
    reviewer_user_id = body.reviewer_user_id
    if not reviewer_user_id and body.reviewer_email:
        reviewer_user = (
            db.execute(select(AppUser).where(AppUser.email == body.reviewer_email))
            .scalars()
            .first()
        )
        if reviewer_user:
            reviewer_user_id = reviewer_user.id

    scope = ReviewScope(
        wave_id=wave_id,
        name=body.name,
        scope_type=body.scope_type,
        scope_filter=body.scope_filter,
        token=token,
        token_expires_at=expires,
        reviewer_name=body.reviewer_name,
        reviewer_email=body.reviewer_email,
        reviewer_user_id=reviewer_user_id,
        status="pending",
    )
    db.add(scope)
    db.flush()

    # Find proposals for this scope and create review items
    preferred_run_id = (wave.config or {}).get("preferred_run_id")
    if preferred_run_id:
        proposals = (
            db.execute(select(CenterProposal).where(CenterProposal.run_id == preferred_run_id))
            .scalars()
            .all()
        )

        # Filter by scope filter (entity codes AND/OR hierarchy nodes)
        scope_ccodes = body.scope_filter.get("entity_ccodes", [])
        hier_nodes = body.scope_filter.get("hierarchy_nodes", [])

        # If hierarchy_nodes specified, resolve which cost centers are under those nodes
        hier_cctrs: set[str] = set()
        if hier_nodes:
            for node_info in hier_nodes:
                node_name = node_info.get("node_name", "")
                if node_name:
                    leaves = (
                        db.execute(select(HierarchyLeaf).where(HierarchyLeaf.setname == node_name))
                        .scalars()
                        .all()
                    )
                    for lf in leaves:
                        hier_cctrs.add(lf.value)

        for p in proposals:
            cc = db.get(LegacyCostCenter, p.legacy_cc_id)
            if not cc:
                continue
            # Entity filter
            if scope_ccodes and cc.ccode not in scope_ccodes:
                continue
            # Hierarchy filter
            if hier_cctrs and (cc.cctr or "") not in hier_cctrs:
                continue
            db.add(
                ReviewItem(
                    scope_id=scope.id,
                    proposal_id=p.id,
                    decision="PENDING",
                )
            )

    # Auto-transition wave through the pipeline
    if wave.status == "proposed":
        wave.status = "in_review"
        wave.locked_at = datetime.now(UTC)
    elif wave.status == "locked":
        wave.status = "in_review"

    db.commit()
    db.refresh(scope)

    # Send invitation email if reviewer_email is provided
    if body.reviewer_email:
        _send_scope_invitation(scope, wave, db)

    return {
        "id": scope.id,
        "name": scope.name,
        "token": token,
        "review_url": f"/review/{token}",
        "status": scope.status,
    }


@router.get("/{wave_id}/scopes")
def list_review_scopes(
    wave_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    scopes = db.execute(select(ReviewScope).where(ReviewScope.wave_id == wave_id)).scalars().all()
    items = []
    for s in scopes:
        total_items = (
            db.execute(
                select(func.count(ReviewItem.id)).where(ReviewItem.scope_id == s.id)
            ).scalar()
            or 0
        )
        decided_items = (
            db.execute(
                select(func.count(ReviewItem.id)).where(
                    ReviewItem.scope_id == s.id,
                    ReviewItem.decision != "PENDING",
                )
            ).scalar()
            or 0
        )
        items.append(
            {
                "id": s.id,
                "name": s.name,
                "scope_type": s.scope_type,
                "status": s.status,
                "reviewer_name": s.reviewer_name,
                "reviewer_email": s.reviewer_email,
                "total_items": total_items,
                "decided_items": decided_items,
                "token_hint": s.token[:8] + "..." if s.token else None,
                "token": s.token,
            }
        )
    return {"wave_id": wave_id, "items": items}


@router.post("/scopes/{scope_id}/invite")
def invite_reviewer(
    scope_id: int,
    db: Session = Depends(get_db),
    _user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    scope = db.get(ReviewScope, scope_id)
    if not scope:
        raise HTTPException(status_code=404, detail="Scope not found")
    if not scope.reviewer_email:
        raise HTTPException(status_code=400, detail="No reviewer email set on this scope")
    if scope.status in ("completed", "revoked", "expired"):
        raise HTTPException(
            status_code=409, detail=f"Cannot invite for scope in status {scope.status}"
        )
    wave = db.get(Wave, scope.wave_id)
    _send_scope_invitation(scope, wave, db)
    scope.status = "invited"
    scope.invited_at = datetime.now(UTC)
    db.commit()
    return {"status": "invited", "scope_id": scope_id}


@router.post("/scopes/{scope_id}/remind")
def remind_reviewer(
    scope_id: int,
    db: Session = Depends(get_db),
    _user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    scope = db.get(ReviewScope, scope_id)
    if not scope:
        raise HTTPException(status_code=404, detail="Scope not found")
    if not scope.reviewer_email:
        raise HTTPException(status_code=400, detail="No reviewer email set on this scope")
    if scope.status in ("completed", "revoked", "expired"):
        raise HTTPException(
            status_code=409,
            detail=f"Cannot send reminder for scope in status {scope.status}",
        )
    wave = db.get(Wave, scope.wave_id)
    _send_scope_invitation(scope, wave, db, is_reminder=True)
    return {"status": "reminder_sent", "scope_id": scope_id}


@router.delete("/scopes/{scope_id}")
def delete_review_scope(
    scope_id: int,
    db: Session = Depends(get_db),
    _user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    scope = db.get(ReviewScope, scope_id)
    if not scope:
        raise HTTPException(status_code=404, detail="Scope not found")
    if scope.status in ("completed", "signed_off"):
        raise HTTPException(
            status_code=409,
            detail=f"Cannot delete scope in status {scope.status}",
        )
    db.delete(scope)
    db.commit()
    return {"deleted": True, "id": scope_id}


class AssignReviewerBody(BaseModel):
    reviewer_name: str | None = None
    reviewer_email: str | None = None
    reviewer_user_id: int | None = None


@router.patch("/scopes/{scope_id}/reviewer")
def assign_reviewer_to_scope(
    scope_id: int,
    body: AssignReviewerBody,
    db: Session = Depends(get_db),
    _user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    """Assign or update the reviewer on an existing scope."""
    scope = db.get(ReviewScope, scope_id)
    if not scope:
        raise HTTPException(status_code=404, detail="Scope not found")
    if scope.status in ("completed", "signed_off"):
        raise HTTPException(
            status_code=409,
            detail=f"Cannot change reviewer on scope in status {scope.status}",
        )
    scope.reviewer_name = body.reviewer_name
    scope.reviewer_email = body.reviewer_email
    # Prefer explicit user ID, then resolve from email
    if body.reviewer_user_id:
        scope.reviewer_user_id = body.reviewer_user_id
    elif body.reviewer_email:
        reviewer_user = (
            db.execute(select(AppUser).where(AppUser.email == body.reviewer_email))
            .scalars()
            .first()
        )
        if reviewer_user:
            scope.reviewer_user_id = reviewer_user.id
        else:
            scope.reviewer_user_id = None
    else:
        scope.reviewer_user_id = None
    db.commit()
    db.refresh(scope)
    return {
        "id": scope.id,
        "reviewer_name": scope.reviewer_name,
        "reviewer_email": scope.reviewer_email,
        "reviewer_user_id": scope.reviewer_user_id,
    }


@router.post("/{wave_id}/proposals/{proposal_id}/override")
def override_proposal(
    wave_id: int,
    proposal_id: int,
    body: dict,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    """Override a proposal's cleansing outcome and/or target object."""
    from app.domain.proposal.service import override_proposal as do_override

    try:
        proposal = do_override(
            proposal_id=proposal_id,
            new_outcome=body.get("outcome", ""),
            new_target=body.get("target_object"),
            reason=body.get("reason", ""),
            user_id=user.id,
            db=db,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from None

    return {
        "id": proposal.id,
        "cleansing_outcome": proposal.cleansing_outcome,
        "target_object": proposal.target_object,
        "override_outcome": proposal.override_outcome,
        "override_target": proposal.override_target,
        "override_reason": proposal.override_reason,
    }


@router.post("/{wave_id}/lock-and-create-targets")
def lock_and_create_targets(
    wave_id: int,
    body: dict,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    """Lock proposals and create target CC/PC objects from approved proposals."""
    from app.domain.proposal.service import lock_proposals

    run_id = body.get("run_id")
    if not run_id:
        # Try to use preferred run from wave config
        wave = db.get(Wave, wave_id)
        if not wave:
            raise HTTPException(status_code=404, detail="Wave not found")
        run_id = (wave.config or {}).get("preferred_run_id")
        if not run_id:
            raise HTTPException(status_code=400, detail="No run_id specified and no preferred run")

    try:
        result = lock_proposals(wave_id, run_id, db)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from None

    return result


@router.post("/{wave_id}/generate-targets")
def generate_targets(
    wave_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    """Generate target CC/PC records from all approved proposals.

    Every KEEP/MERGE/REDESIGN proposal gets a new target CC (and PC if
    CC_AND_PC). All SAP attributes are inherited from the legacy source.
    RETIRE proposals get a mapping record but no target object.
    """
    from app.domain.proposal.service import generate_wave_targets

    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.status not in ("signed_off", "closed", "locked", "in_review"):
        raise HTTPException(
            status_code=409,
            detail=f"Wave must be signed_off/closed for target generation (current: {wave.status})",
        )

    try:
        result = generate_wave_targets(wave_id, db)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from None

    return result


@router.get("/{wave_id}/target-summary")
def target_summary(
    wave_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    """Get summary of generated targets for a wave."""
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")

    cc_count = (
        db.execute(
            select(func.count(TargetCostCenter.id)).where(
                TargetCostCenter.approved_in_wave == wave_id
            )
        ).scalar()
        or 0
    )
    pc_count = (
        db.execute(
            select(func.count(TargetProfitCenter.id)).where(
                TargetProfitCenter.approved_in_wave == wave_id
            )
        ).scalar()
        or 0
    )
    mapping_count = (
        db.execute(
            select(func.count(CenterMapping.id)).where(
                CenterMapping.notes.like(f"%wave:{wave_id}%")
            )
        ).scalar()
        or 0
    )

    # Breakdown by mapping type
    type_counts = {}
    for mapping_type in ["1:1", "merge", "redesign", "retire"]:
        cnt = (
            db.execute(
                select(func.count(CenterMapping.id)).where(
                    CenterMapping.notes.like(f"%wave:{wave_id}%"),
                    CenterMapping.mapping_type == mapping_type,
                )
            ).scalar()
            or 0
        )
        type_counts[mapping_type] = cnt

    return {
        "wave_id": wave_id,
        "target_cc_count": cc_count,
        "target_pc_count": pc_count,
        "mapping_count": mapping_count,
        "mapping_breakdown": type_counts,
        "has_targets": cc_count > 0 or pc_count > 0,
    }


@router.get("/{wave_id}/mdg-export")
def mdg_export(
    wave_id: int,
    export_type: str = "cost_center",
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    """Generate MDG export file for the wave's target objects.

    Export types:
    - cost_center: Target CC list (SAP CSKS structure) with full SAP fields
    - profit_center: Target PC list (SAP CEPC structure) with full SAP fields
    - retire: Decommission list for legacy centers to deactivate
    - mapping: Old→New CC/PC mapping table
    """
    from app.infra.mdg.export import (
        export_cost_centers,
        export_mapping_table,
        export_profit_centers,
        export_retire_list,
    )

    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.status not in ("signed_off", "closed"):
        raise HTTPException(status_code=409, detail="Wave must be signed_off or closed for export")

    if export_type == "cost_center":
        targets = (
            db.execute(select(TargetCostCenter).where(TargetCostCenter.approved_in_wave == wave_id))
            .scalars()
            .all()
        )
        if not targets:
            raise HTTPException(
                status_code=404,
                detail="No target cost centers found. Run 'Generate Target Numbers' first.",
            )
        centers = [_cc_to_export_dict(t) for t in targets]
        result = export_cost_centers(centers, wave_id)

    elif export_type == "profit_center":
        targets = (
            db.execute(
                select(TargetProfitCenter).where(TargetProfitCenter.approved_in_wave == wave_id)
            )
            .scalars()
            .all()
        )
        if not targets:
            raise HTTPException(
                status_code=404,
                detail="No target profit centers found. Run 'Generate Target Numbers' first.",
            )
        centers = [_pc_to_export_dict(t) for t in targets]
        result = export_profit_centers(centers, wave_id)

    elif export_type == "retire":
        preferred_run_id = (wave.config or {}).get("preferred_run_id")
        if not preferred_run_id:
            raise HTTPException(status_code=400, detail="No preferred run set")
        proposals = (
            db.execute(
                select(CenterProposal).where(
                    CenterProposal.run_id == preferred_run_id,
                    CenterProposal.cleansing_outcome == "RETIRE",
                )
            )
            .scalars()
            .all()
        )
        centers = []
        for p in proposals:
            cc = db.get(LegacyCostCenter, p.legacy_cc_id)
            if cc:
                centers.append(_legacy_cc_to_export_dict(cc))
        result = export_retire_list(centers, wave_id)

    elif export_type == "mapping":
        mappings = (
            db.execute(select(CenterMapping).where(CenterMapping.notes.like(f"%wave:{wave_id}%")))
            .scalars()
            .all()
        )
        if not mappings:
            raise HTTPException(
                status_code=404,
                detail="No mapping records found. Run 'Generate Target Numbers' first.",
            )
        result = export_mapping_table(mappings, wave_id)

    else:
        raise HTTPException(status_code=400, detail=f"Invalid export_type: {export_type}")

    return {
        "filename": result.filename,
        "content": result.content,
        "record_count": result.record_count,
        "export_type": result.export_type,
    }


def _cc_to_export_dict(t: TargetCostCenter) -> dict:
    """Convert a TargetCostCenter row to a dict for MDG export with full SAP fields."""
    return {
        "cctr": t.cctr,
        "coarea": t.coarea,
        "txtsh": t.txtsh,
        "txtmi": t.txtmi,
        "responsible": t.responsible,
        "ccode": t.ccode,
        "cctrcgy": t.cctrcgy,
        "currency": t.currency,
        "pctr": t.pctr,
        "gsber": t.gsber,
        "werks": t.werks,
        "kalsm": t.kalsm,
        "txjcd": t.txjcd,
        "func_area": t.func_area,
        "land1": t.land1,
        "name1": t.name1,
        "regio": t.regio,
        "abtei": t.abtei,
        "datab": t.datab,
        "datbi": t.datbi,
    }


def _pc_to_export_dict(t: TargetProfitCenter) -> dict:
    """Convert a TargetProfitCenter row to a dict for MDG export with full SAP fields."""
    return {
        "pctr": t.pctr,
        "coarea": t.coarea,
        "txtsh": t.txtsh,
        "txtmi": t.txtmi,
        "responsible": t.responsible,
        "ccode": t.ccode,
        "currency": t.currency,
        "department": t.department,
        "land1": t.land1,
        "name1": t.name1,
        "regio": t.regio,
        "datab": t.datab,
        "datbi": t.datbi,
        "segment": t.segment,
    }


def _legacy_cc_to_export_dict(cc: LegacyCostCenter) -> dict:
    """Convert a LegacyCostCenter to a dict for the retire export."""
    return {
        "cctr": cc.cctr,
        "coarea": cc.coarea,
        "txtsh": cc.txtsh,
        "txtmi": cc.txtmi,
        "responsible": cc.responsible,
        "ccode": cc.ccode,
        "cctrcgy": cc.cctrcgy,
        "currency": cc.currency,
        "pctr": cc.pctr,
    }


@router.get("/{wave_id}/progress")
def wave_progress(wave_id: int, db: Session = Depends(get_db)) -> dict:
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    scopes = db.execute(select(ReviewScope).where(ReviewScope.wave_id == wave.id)).scalars().all()
    total_review_items = 0
    decided_items = 0
    for s in scopes:
        t = (
            db.execute(
                select(func.count(ReviewItem.id)).where(ReviewItem.scope_id == s.id)
            ).scalar()
            or 0
        )
        d = (
            db.execute(
                select(func.count(ReviewItem.id)).where(
                    ReviewItem.scope_id == s.id, ReviewItem.decision != "PENDING"
                )
            ).scalar()
            or 0
        )
        total_review_items += t
        decided_items += d
    # Auto-correct wave status based on actual review state
    status_changed = False
    if scopes and wave.status in ("proposed", "locked"):
        # Scopes exist → should be at least in_review
        wave.status = "in_review"
        if not wave.locked_at:
            wave.locked_at = datetime.now(UTC)
        status_changed = True
    if wave.status == "in_review" and scopes:
        all_complete = all(s.status == "completed" for s in scopes)
        if all_complete and total_review_items > 0 and decided_items >= total_review_items:
            wave.status = "signed_off"
            wave.signed_off_at = datetime.now(UTC)
            status_changed = True
    if status_changed:
        db.commit()

    return {
        "wave_id": wave.id,
        "status": wave.status,
        "total_review_items": total_review_items,
        "decided_items": decided_items,
        "completion_pct": (
            round(decided_items / total_review_items * 100, 1) if total_review_items else 0
        ),
        "scopes": [
            {"id": s.id, "name": s.name, "status": s.status, "scope_type": s.scope_type}
            for s in scopes
        ],
    }


# --- Reviewer Workload Balancer ---


@router.get("/{wave_id}/workload")
def reviewer_workload(
    wave_id: int,
    db: Session = Depends(get_db),
    _user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    """Get reviewer workload distribution for a wave."""
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")

    scopes = db.execute(select(ReviewScope).where(ReviewScope.wave_id == wave_id)).scalars().all()

    workload = []
    for s in scopes:
        total_items = (
            db.execute(
                select(func.count(ReviewItem.id)).where(ReviewItem.scope_id == s.id)
            ).scalar()
            or 0
        )
        decided = (
            db.execute(
                select(func.count(ReviewItem.id)).where(
                    ReviewItem.scope_id == s.id,
                    ReviewItem.decision != "PENDING",
                )
            ).scalar()
            or 0
        )
        pending = total_items - decided
        workload.append(
            {
                "scope_id": s.id,
                "name": s.name,
                "reviewer": s.reviewer_email,
                "status": s.status,
                "total_items": total_items,
                "decided": decided,
                "pending": pending,
                "completion_pct": round(decided / total_items * 100, 1) if total_items else 0,
            }
        )

    total_all = sum(w["total_items"] for w in workload)
    avg_per_reviewer = round(total_all / len(workload), 1) if workload else 0
    max_load = max((w["total_items"] for w in workload), default=0)
    min_load = min((w["total_items"] for w in workload), default=0)
    imbalance = round((max_load - min_load) / avg_per_reviewer * 100, 1) if avg_per_reviewer else 0

    return {
        "wave_id": wave_id,
        "reviewers": workload,
        "summary": {
            "total_items": total_all,
            "total_reviewers": len(workload),
            "avg_per_reviewer": avg_per_reviewer,
            "max_load": max_load,
            "min_load": min_load,
            "imbalance_pct": imbalance,
        },
    }


# --- Auto-Approve ---


class AutoApproveParams(BaseModel):
    confidence_threshold: float = 0.85
    max_items: int | None = None
    verdicts: list[str] = ["KEEP"]


@router.post("/{wave_id}/auto-approve")
def auto_approve_obvious(
    wave_id: int,
    params: AutoApproveParams,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    """Auto-approve review items where the analysis confidence is above threshold."""
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.status != "in_review":
        raise HTTPException(status_code=409, detail="Wave must be in_review for auto-approve")

    preferred_run_id = (wave.config or {}).get("preferred_run_id")
    if not preferred_run_id:
        raise HTTPException(status_code=409, detail="No preferred run set")

    # Get pending review items with their proposals
    pending = db.execute(
        select(ReviewItem, CenterProposal)
        .join(CenterProposal, ReviewItem.proposal_id == CenterProposal.id)
        .join(ReviewScope, ReviewItem.scope_id == ReviewScope.id)
        .where(
            ReviewScope.wave_id == wave_id,
            ReviewItem.decision == "PENDING",
            CenterProposal.run_id == preferred_run_id,
        )
    ).all()

    approved_count = 0
    for item, proposal in pending:
        confidence = float(proposal.confidence or 0)
        verdict_match = proposal.cleansing_outcome in params.verdicts
        if verdict_match and confidence >= params.confidence_threshold:
            item.decision = "APPROVED"
            item.decided_by = f"auto:{user.id}"
            item.decided_at = datetime.now(UTC)
            approved_count += 1
            if params.max_items and approved_count >= params.max_items:
                break

    db.commit()
    return {
        "approved": approved_count,
        "remaining_pending": len(pending) - approved_count,
        "threshold": params.confidence_threshold,
        "verdicts": params.verdicts,
    }


# --- Workload-Aware Scope Assignment ---


class AutoAssignParams(BaseModel):
    reviewer_emails: list[str]
    scope_type: str = "entity"
    strategy: str = "round_robin"  # round_robin | balanced | entity_group


@router.post("/{wave_id}/auto-assign")
def auto_assign_scopes(
    wave_id: int,
    params: AutoAssignParams,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin")),
) -> dict:
    """Automatically create review scopes and distribute proposals to reviewers."""
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.status not in ("locked", "in_review"):
        raise HTTPException(status_code=409, detail="Wave must be locked or in_review")

    preferred_run_id = (wave.config or {}).get("preferred_run_id")
    if not preferred_run_id:
        raise HTTPException(status_code=409, detail="No preferred run set")

    proposals = (
        db.execute(select(CenterProposal).where(CenterProposal.run_id == preferred_run_id))
        .scalars()
        .all()
    )
    if not proposals:
        raise HTTPException(status_code=409, detail="No proposals found for preferred run")

    reviewers = list(dict.fromkeys(params.reviewer_emails))  # deduplicate preserving order
    if not reviewers:
        raise HTTPException(status_code=400, detail="At least one reviewer required")

    # Group proposals by entity if strategy is entity_group
    if params.strategy == "entity_group":
        entity_groups: dict[str, list] = {}
        for p in proposals:
            cc = db.get(LegacyCostCenter, p.legacy_cc_id)
            ccode = cc.ccode if cc else "UNKNOWN"
            entity_groups.setdefault(ccode, []).append(p)
        # Distribute entity groups to reviewers round-robin
        sorted_groups = sorted(entity_groups.items(), key=lambda x: -len(x[1]))
        reviewer_loads: dict[str, list] = {r: [] for r in reviewers}
        for _ccode, group_proposals in sorted_groups:
            min_reviewer = min(reviewer_loads, key=lambda r: len(reviewer_loads[r]))
            reviewer_loads[min_reviewer].extend(group_proposals)
    else:
        # round_robin or balanced
        reviewer_loads = {r: [] for r in reviewers}
        for i, p in enumerate(proposals):
            target = reviewers[i % len(reviewers)]
            reviewer_loads[target].append(p)

    created_scopes = []
    for email, assigned_proposals in reviewer_loads.items():
        if not assigned_proposals:
            continue
        token = secrets.token_urlsafe(32)
        expires = datetime.now(UTC) + timedelta(days=30)
        scope = ReviewScope(
            wave_id=wave_id,
            name=f"Auto-assigned: {email}",
            scope_type=params.scope_type,
            scope_filter={"auto_assigned": True, "reviewer": email},
            token=token,
            token_expires_at=expires,
            reviewer_email=email,
            status="pending",
        )
        db.add(scope)
        db.flush()
        for p in assigned_proposals:
            db.add(ReviewItem(scope_id=scope.id, proposal_id=p.id, decision="PENDING"))
        created_scopes.append(
            {
                "scope_id": scope.id,
                "reviewer": email,
                "items": len(assigned_proposals),
                "review_url": f"/review/{token}",
            }
        )

    if wave.status == "locked":
        wave.status = "in_review"

    db.commit()
    return {
        "scopes_created": len(created_scopes),
        "scopes": created_scopes,
        "total_items": len(proposals),
        "strategy": params.strategy,
    }


# ── V2 CEMA Migration endpoints ─────────────────────────────────────────


class V2AnalysisParams(BaseModel):
    config_id: int | None = None
    pc_approach_rules: list[dict] | None = None
    pc_start: int = 137
    cc_start: int = 1
    mode: str = "simulation"  # simulation | activated
    label: str | None = None
    excluded_scopes: list[int] | None = None


def _run_v2_in_thread(
    run_id: int,
    wave_id: int | None,
    config_id: int,
    user_id: int,
    *,
    mode: str,
    id_config: dict | None,
) -> None:
    """Daemon-thread runner for the V2 CEMA migration pipeline.

    Mirrors :func:`app.api.runs._run_global_in_thread` (the V1 dispatch
    helper) so the V2 endpoint can return immediately while the long
    pipeline runs in the background. Without this the request handler
    blocked for the full duration of the analysis (minutes on real
    data), the frontend had no progress to show, and a request timeout
    would leave the run in 'running' status with no way to recover.

    The same trade-offs apply: if the API process restarts mid-run the
    thread dies and the row is left in 'running' — no different from
    the prior synchronous behaviour. Operators who need durability
    should switch to the Celery dispatch path
    (``app.workers.tasks.run_v2_analysis``) which is already wired up
    but unused in single-process deployments.
    """
    from app.infra.db.session import SessionLocal
    from app.services.analysis_v2 import execute_v2_analysis_for_run

    db = SessionLocal()
    try:
        run = db.get(AnalysisRun, run_id)
        if not run:
            log.warning("v2_run_thread.not_found run_id=%s", run_id)
            return
        execute_v2_analysis_for_run(
            run=run,
            wave_id=wave_id,
            config_id=config_id,
            mode=mode,
            id_config=id_config,
            db=db,
        )
    except Exception:
        log.exception("v2_run_thread.failed run_id=%s", run_id)
        # Mark the run as failed so the UI can surface the error
        # instead of leaving it stuck in 'running'.
        try:
            run = db.get(AnalysisRun, run_id)
            if run and run.status not in ("completed", "cancelled"):
                run.status = "failed"
                run.finished_at = datetime.now(UTC)
                db.commit()
        except Exception:
            log.exception("v2_run_thread.failed_to_mark_failed run_id=%s", run_id)
    finally:
        db.close()


@router.post("/{wave_id}/analyse-v2")
def run_v2_analysis_endpoint(
    wave_id: int,
    params: V2AnalysisParams | None = None,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    """Queue a V2 CEMA migration run on the wave's cost centers.

    Returns immediately with the run id; the actual pipeline executes
    in a daemon thread. Frontend polls ``GET /api/runs/{id}`` every 2s
    to show progress and re-renders when ``status`` becomes terminal.
    Replaces the previous synchronous behaviour where the request hung
    for the full duration of the pipeline.
    """
    from app.services.analysis_v2 import V2_DEFAULT_CONFIG

    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.status not in ("draft", "analysing", "proposed"):
        raise HTTPException(status_code=409, detail=f"Cannot analyse wave in status {wave.status}")

    # If inline params provided, build a runtime config
    config_id = None
    if params and params.config_id:
        config_id = params.config_id
    elif params and params.pc_approach_rules is not None:
        import copy

        from app.models.core import AnalysisConfig

        cfg = copy.deepcopy(V2_DEFAULT_CONFIG)
        for step in cfg["pipeline"]:
            if step["routine"] == "v2.pc_approach":
                step["params"]["approach_rules"] = params.pc_approach_rules
        cfg["id_assignment"]["pc_start"] = params.pc_start
        cfg["id_assignment"]["cc_start"] = params.cc_start
        max_ver = (
            db.execute(
                select(func.coalesce(func.max(AnalysisConfig.version), 0)).where(
                    AnalysisConfig.code == "cema_migration_v2"
                )
            ).scalar()
            or 0
        )
        ac = AnalysisConfig(
            code="cema_migration_v2",
            version=max_ver + 1,
            name="V2 CEMA Migration (runtime)",
            config=cfg,
            created_by=user.id,
        )
        db.add(ac)
        db.flush()
        config_id = ac.id

    if config_id is None:
        # Fall back to the latest persisted V2 config so callers don't
        # have to re-specify pc_approach_rules every time.
        from app.models.core import AnalysisConfig

        latest = db.execute(
            select(AnalysisConfig)
            .where(AnalysisConfig.code == "cema_migration_v2")
            .order_by(AnalysisConfig.version.desc())
            .limit(1)
        ).scalar_one_or_none()
        if latest:
            config_id = latest.id
        else:
            # No persisted config — create a default one inline so the
            # FK on AnalysisRun.config_id can be satisfied.
            ac = AnalysisConfig(
                code="cema_migration_v2",
                version=1,
                name="V2 CEMA Migration (default)",
                config=V2_DEFAULT_CONFIG,
                created_by=user.id,
            )
            db.add(ac)
            db.flush()
            config_id = ac.id

    sim_mode = params.mode if params else "simulation"
    sim_label = params.label if params else None
    excl = [str(x) for x in (params.excluded_scopes or [])] if params else None
    id_config = {
        "pc_start": params.pc_start if params else 137,
        "cc_start": params.cc_start if params else 1,
    }

    # Create the run row in 'queued' status so the UI immediately sees
    # something to render and so the per-CC progress fields exist
    # before the thread starts updating them.
    run = AnalysisRun(
        wave_id=wave_id,
        mode=sim_mode,
        status="queued",
        engine_version="v2.cema_migration",
        config_id=config_id,
        label=sim_label,
        excluded_scopes=excl,
        triggered_by=user.id,
        started_at=datetime.now(UTC),
        total_centers=0,
        completed_centers=0,
    )
    db.add(run)
    db.commit()
    db.refresh(run)

    thread = threading.Thread(
        target=_run_v2_in_thread,
        args=(run.id, wave_id, config_id, user.id),
        kwargs={"mode": sim_mode, "id_config": id_config},
        daemon=True,
        name=f"v2_run_{run.id}",
    )
    thread.start()

    return {
        "run_id": run.id,
        "status": run.status,
        "started_at": str(run.started_at) if run.started_at else None,
    }


@router.get("/{wave_id}/runs/{run_id}/export-v2")
def export_v2_results(
    wave_id: int,
    run_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> Response:
    """Export V2 analysis results as Excel with PC/CC templates + mapping."""
    import io

    from openpyxl import Workbook

    run = db.get(AnalysisRun, run_id)
    if not run or run.wave_id != wave_id:
        raise HTTPException(status_code=404, detail="Run not found for this wave")

    proposals = (
        db.execute(
            select(CenterProposal)
            .where(CenterProposal.run_id == run_id)
            .order_by(CenterProposal.id)
        )
        .scalars()
        .all()
    )

    wb = Workbook()

    # Sheet 1: PC Template
    ws_pc = wb.active
    ws_pc.title = "PC_Template"  # type: ignore[union-attr]
    pc_headers = [
        "PC_ID",
        "PC_Name",
        "CO_Area",
        "Company_Code",
        "Currency",
        "Responsible",
        "Approach",
        "Group_Key",
        "Migrate",
    ]
    ws_pc.append(pc_headers)  # type: ignore[union-attr]

    # Sheet 2: CC Template
    ws_cc = wb.create_sheet("CC_Template")
    cc_headers = [
        "CC_ID",
        "CC_Name",
        "CO_Area",
        "Company_Code",
        "Currency",
        "Responsible",
        "Category",
        "Legacy_CCTR",
        "Migrate",
    ]
    ws_cc.append(cc_headers)

    # Sheet 3: Mapping
    ws_map = wb.create_sheet("Mapping")
    map_headers = [
        "Legacy_CCTR",
        "Legacy_Name",
        "CO_Area",
        "Company_Code",
        "Migrate",
        "Approach",
        "PC_ID",
        "PC_Name",
        "CC_ID",
        "CC_Name",
        "External_Hierarchy",
        "CEMA_Hierarchy",
    ]
    ws_map.append(map_headers)

    # Track PCs already written (for 1:n dedup)
    pc_written: set[str] = set()

    for p in proposals:
        attrs = p.attrs or {}
        legacy = db.get(LegacyCostCenter, p.legacy_cc_id)
        if not legacy:
            continue

        migrate = attrs.get("migrate", "N")
        approach = attrs.get("approach", "1:1")
        pc_id = attrs.get("pc_id", "")
        pc_name = attrs.get("pc_name", "")
        cc_id = attrs.get("cc_id", "")
        cc_name = attrs.get("cc_name", legacy.txtsh or "")
        group_key = attrs.get("group_key", "")
        ext_hierarchy = attrs.get("ext_hierarchy", "")
        cema_hierarchy = attrs.get("cema_hierarchy", "")

        # Mapping row — all centers
        ws_map.append(
            [
                legacy.cctr,
                legacy.txtsh,
                legacy.coarea,
                legacy.ccode,
                migrate,
                approach,
                pc_id,
                pc_name,
                cc_id,
                cc_name,
                ext_hierarchy,
                cema_hierarchy,
            ]
        )

        if migrate != "Y":
            continue

        # PC Template (deduplicated for 1:n)
        if pc_id and pc_id not in pc_written:
            ws_pc.append(
                [  # type: ignore[union-attr]
                    pc_id,
                    pc_name,
                    legacy.coarea,
                    legacy.ccode,
                    legacy.currency,
                    legacy.responsible,
                    approach,
                    group_key,
                    "Y",
                ]
            )
            pc_written.add(pc_id)

        # CC Template
        if cc_id:
            ws_cc.append(
                [
                    cc_id,
                    cc_name,
                    legacy.coarea,
                    legacy.ccode,
                    legacy.currency,
                    legacy.responsible,
                    legacy.cctrcgy,
                    legacy.cctr,
                    "Y",
                ]
            )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"v2_migration_wave{wave_id}_run{run_id}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{wave_id}/runs/{run_id}/proposals-v2")
def list_v2_proposals(
    wave_id: int,
    run_id: int,
    migrate: str | None = None,
    approach: str | None = None,
    page: int = 1,
    per_page: int = 50,
    include_paths: bool = False,
    hierarchy_id: int | None = None,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager", "viewer")),
) -> dict:
    """List V2 proposals with migration details.

    When ``include_paths=true`` is set, each item also gets a
    ``hierarchy_path`` array (the chain of node setnames from root to
    leaf) and the response carries a top-level ``hierarchy_max_depth``
    so the frontend knows how many ``L0..LX`` columns to render.

    The hierarchy used for path resolution is:
    * the hierarchy passed via ``hierarchy_id`` (explicit),
    * or the wave's first ``WaveHierarchyScope`` row (most relevant
      to the operator who scoped the wave by hierarchy),
    * or the first active CC hierarchy (setclass=0101),
    * or no path resolution at all (empty paths) if none exist.

    The default is ``include_paths=false`` so existing callers don't
    pay the resolution cost; the wave-detail simulation view opts in.
    """
    run = db.get(AnalysisRun, run_id)
    if not run or run.wave_id != wave_id:
        raise HTTPException(status_code=404, detail="Run not found for this wave")

    q = select(CenterProposal).where(CenterProposal.run_id == run_id)

    # Apply JSONB filters at SQL level
    if migrate:
        q = q.where(CenterProposal.attrs["migrate"].astext == migrate)
    if approach:
        q = q.where(CenterProposal.attrs["approach"].astext == approach)

    # Count total (after filters)
    count_q = select(func.count()).select_from(q.subquery())
    total = db.execute(count_q).scalar() or 0

    # Paginate
    proposals = (
        db.execute(q.order_by(CenterProposal.id).offset((page - 1) * per_page).limit(per_page))
        .scalars()
        .all()
    )

    items = []
    page_cctrs: list[str] = []
    for p in proposals:
        attrs = p.attrs or {}
        legacy = db.get(LegacyCostCenter, p.legacy_cc_id)
        cctr = legacy.cctr if legacy else ""
        if cctr:
            page_cctrs.append(cctr)
        items.append(
            {
                "id": p.id,
                "legacy_cctr": cctr,
                "legacy_name": legacy.txtsh if legacy else "",
                "coarea": legacy.coarea if legacy else "",
                "ccode": legacy.ccode if legacy else "",
                "migrate": attrs.get("migrate", "N"),
                "approach": attrs.get("approach", "1:1"),
                "pc_id": attrs.get("pc_id", ""),
                "pc_name": attrs.get("pc_name", ""),
                "cc_id": attrs.get("cc_id", ""),
                "cc_name": attrs.get("cc_name", ""),
                "cleansing_outcome": p.cleansing_outcome,
                "rule_path": p.rule_path,
            }
        )

    response: dict = {
        "total": total,
        "page": page,
        "per_page": per_page,
        "items": items,
    }

    if include_paths and page_cctrs:
        from app.api.reference import _resolve_hierarchy_paths
        from app.models.core import Hierarchy

        resolved_hier_id = hierarchy_id
        if resolved_hier_id is None:
            # Prefer a hierarchy the operator explicitly scoped this wave by
            wh_row = (
                db.execute(
                    select(WaveHierarchyScope).where(WaveHierarchyScope.wave_id == wave_id).limit(1)
                )
                .scalars()
                .first()
            )
            if wh_row is not None:
                resolved_hier_id = wh_row.hierarchy_id
        if resolved_hier_id is None:
            # Fallback: first active CC hierarchy
            cc_hier = (
                db.execute(
                    select(Hierarchy)
                    .where(Hierarchy.is_active.is_(True))
                    .where(Hierarchy.setclass.in_(["0101", "CC"]))
                    .order_by(Hierarchy.id)
                    .limit(1)
                )
                .scalars()
                .first()
            )
            if cc_hier is not None:
                resolved_hier_id = cc_hier.id

        if resolved_hier_id is not None:
            paths, max_depth = _resolve_hierarchy_paths(db, resolved_hier_id, page_cctrs)
            for it in items:
                it["hierarchy_path"] = paths.get(it["legacy_cctr"], [])
            response["hierarchy_id"] = resolved_hier_id
            response["hierarchy_max_depth"] = max_depth
        else:
            response["hierarchy_id"] = None
            response["hierarchy_max_depth"] = 0

    return response


# ── Global simulation endpoint ───────────────────────────────────────────


class GlobalSimParams(BaseModel):
    config_id: int | None = None
    pc_approach_rules: list[dict] | None = None
    pc_start: int = 137
    cc_start: int = 1
    mode: str = "simulation"
    label: str | None = None
    excluded_scopes: list[int] | None = None


@router.post("/global/simulate-v2")
def run_global_v2_simulation(
    params: GlobalSimParams | None = None,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    """Run V2 decision tree on ALL centers (global mode).

    Optionally exclude centers belonging to specific wave IDs.
    """
    from app.services.analysis_v2 import V2_DEFAULT_CONFIG, run_v2_analysis

    config_id = None
    if params and params.config_id:
        config_id = params.config_id
    elif params and params.pc_approach_rules is not None:
        import copy

        from app.models.core import AnalysisConfig

        cfg = copy.deepcopy(V2_DEFAULT_CONFIG)
        for step in cfg["pipeline"]:
            if step["routine"] == "v2.pc_approach":
                step["params"]["approach_rules"] = params.pc_approach_rules
        cfg["id_assignment"]["pc_start"] = params.pc_start
        cfg["id_assignment"]["cc_start"] = params.cc_start
        max_ver = (
            db.execute(
                select(func.coalesce(func.max(AnalysisConfig.version), 0)).where(
                    AnalysisConfig.code == "cema_migration_v2"
                )
            ).scalar()
            or 0
        )
        ac = AnalysisConfig(
            code="cema_migration_v2",
            version=max_ver + 1,
            name="V2 CEMA Migration (global runtime)",
            config=cfg,
            created_by=user.id,
        )
        db.add(ac)
        db.flush()
        config_id = ac.id

    sim_mode = params.mode if params else "simulation"
    sim_label = params.label if params else None
    excl = [str(x) for x in (params.excluded_scopes or [])] if params else None

    try:
        result = run_v2_analysis(
            None,
            config_id,
            db,
            user.id,
            mode=sim_mode,
            label=sim_label,
            excluded_scopes=excl,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"V2 global simulation failed: {e}") from None

    return result


# ── Simulation management ────────────────────────────────────────────────


@router.get("/simulations/v2")
def list_simulations(
    wave_id: int | None = None,
    mode: str | None = None,
    engine: str | None = None,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager", "viewer")),
) -> dict:
    """List simulation/activated runs (V1 and V2), optionally filtered."""
    q = select(AnalysisRun)
    if wave_id is not None:
        q = q.where(AnalysisRun.wave_id == wave_id)
    if mode:
        q = q.where(AnalysisRun.mode == mode)
    if engine:
        q = q.where(AnalysisRun.engine_version == engine)
    q = q.order_by(AnalysisRun.created_at.desc())

    runs = db.execute(q).scalars().all()
    items = []
    for r in runs:
        items.append(
            {
                "id": r.id,
                "wave_id": r.wave_id,
                "mode": r.mode or "simulation",
                "label": r.label,
                "status": r.status,
                "engine_version": r.engine_version,
                "total_centers": r.total_centers,
                "completed_centers": r.completed_centers,
                "config_id": r.config_id,
                "started_at": r.started_at.isoformat() if r.started_at else None,
                "finished_at": r.finished_at.isoformat() if r.finished_at else None,
                "kpis": r.kpis,
                "excluded_scopes": r.excluded_scopes,
            }
        )
    return {"items": items, "total": len(items)}


@router.delete("/simulations/{run_id}")
def delete_simulation(
    run_id: int,
    db: Session = Depends(get_db),
    _user: AppUser = Depends(require_role("admin")),
) -> dict:
    """Hard-delete a simulation run and all its proposals.

    Children rows in ``center_proposal``, ``llm_call``, ``run_step``, and
    similar tables that reference ``analysis_run.id`` are removed
    automatically by ``ON DELETE CASCADE`` on those FKs (see model
    definitions). Activated runs cannot be deleted — they represent
    committed state and require an explicit revert flow that doesn't
    exist yet.

    Wave status is intentionally NOT changed by this operation. If the
    wave was bumped to ``analysing`` because of this run, the operator
    needs to roll the wave status back manually if appropriate.
    """
    run = db.get(AnalysisRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    if run.mode == "activated":
        raise HTTPException(
            status_code=409,
            detail="Cannot delete an activated run — revert it first",
        )
    db.delete(run)
    db.commit()
    return {"status": "deleted", "id": run_id}


@router.post("/simulations/{run_id}/activate")
def activate_v2_simulation(
    run_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin")),
) -> dict:
    """Activate a simulation — re-assign real PC/CC IDs and mark as activated.

    This creates final P/C IDs (instead of PT/CT temp IDs) and locks the run.
    After activation the data manager can release for review.
    """
    from app.services.analysis_v2 import assign_v2_ids

    run = db.get(AnalysisRun, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    if run.engine_version != "v2.cema_migration":
        raise HTTPException(status_code=400, detail="Not a V2 run")
    if run.mode == "activated":
        raise HTTPException(status_code=409, detail="Run already activated")
    if run.status != "completed":
        raise HTTPException(status_code=409, detail="Can only activate completed runs")

    # Get config for real ID parameters
    from app.models.core import AnalysisConfig as AnalysisCfg

    ac = db.get(AnalysisCfg, run.config_id)
    id_config = (ac.config or {}).get("id_assignment", {}) if ac else {}

    # Re-assign with real prefixes
    id_result = assign_v2_ids(
        run_id=run.id,
        db=db,
        pc_prefix=id_config.get("pc_prefix", "P"),
        cc_prefix=id_config.get("cc_prefix", "C"),
        pc_start=id_config.get("pc_start", 137),
        cc_start=id_config.get("cc_start", 1),
        id_width=id_config.get("id_width", 5),
    )

    run.mode = "activated"
    run.label = (run.label or "V2") + " [ACTIVATED]"

    # Update wave status if wave-scoped and transition is valid
    if run.wave_id:
        wave = db.get(Wave, run.wave_id)
        if wave and "proposed" in VALID_TRANSITIONS.get(wave.status, []):
            wave.status = "proposed"
            wave.preferred_run_id = run.id
            wave.config = {**(wave.config or {}), "preferred_run_id": run.id}

    db.commit()

    return {
        "run_id": run.id,
        "mode": "activated",
        "id_assignment": id_result,
    }


# ── V2 hierarchy nodes (for PC approach picker) ─────────────────────────


@router.get("/hierarchy-nodes")
def list_hierarchy_nodes_for_picker(
    hierarchy_id: int | None = None,
    setclass: str | None = None,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager", "viewer")),
) -> dict:
    """Return hierarchy nodes in a tree structure for the PC approach node picker."""
    from app.models.core import Hierarchy, HierarchyNode

    if hierarchy_id:
        hier = db.get(Hierarchy, hierarchy_id)
        if not hier:
            raise HTTPException(status_code=404, detail="Hierarchy not found")
        nodes = (
            db.execute(
                select(HierarchyNode)
                .where(HierarchyNode.hierarchy_id == hierarchy_id)
                .order_by(HierarchyNode.seq)
            )
            .scalars()
            .all()
        )
    else:
        q = select(Hierarchy).order_by(Hierarchy.setclass, Hierarchy.setname)
        if setclass:
            q = q.where(Hierarchy.setclass == setclass)
        hiers = db.execute(q).scalars().all()
        return {
            "hierarchies": [
                {
                    "id": h.id,
                    "setclass": h.setclass,
                    "setname": h.setname,
                    "label": h.label or h.setname,
                    "scope": h.scope,
                    "data_category": h.data_category,
                }
                for h in hiers
            ]
        }

    # Build tree using setname-based parent-child relationships
    # Each HierarchyNode has parent_setname and child_setname
    children_of: dict[str, list[dict]] = {}
    all_children: set[str] = set()
    all_parents: set[str] = set()

    for n in nodes:
        all_parents.add(n.parent_setname)
        all_children.add(n.child_setname)
        nd = {
            "id": n.id,
            "node_name": n.child_setname,
            "node_text": n.child_setname,
            "parent_name": n.parent_setname,
            "children": [],
        }
        children_of.setdefault(n.parent_setname, []).append(nd)

    # Root parents are those that appear as parent but never as child
    root_parents = all_parents - all_children

    # Build tree recursively
    def _build_subtree(parent_name: str) -> list[dict]:
        result = []
        for nd in children_of.get(parent_name, []):
            nd["children"] = _build_subtree(nd["node_name"])
            result.append(nd)
        return result

    root_nodes: list[dict] = []
    for rp in sorted(root_parents):
        for nd in children_of.get(rp, []):
            nd["children"] = _build_subtree(nd["node_name"])
            root_nodes.append(nd)

    return {"hierarchy_id": hierarchy_id, "nodes": root_nodes, "total_nodes": len(nodes)}


# ── V2 scope coverage dashboard ──────────────────────────────────────────


@router.get("/{wave_id}/scope-coverage")
def get_wave_scope_coverage(
    wave_id: int,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager", "viewer")),
) -> dict:
    """Check whether all centers in a wave are covered by review scopes + reviewers."""
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")

    # Get all proposal IDs for latest V2 run
    latest_run = db.execute(
        select(AnalysisRun)
        .where(AnalysisRun.wave_id == wave_id, AnalysisRun.engine_version == "v2.cema_migration")
        .order_by(AnalysisRun.created_at.desc())
        .limit(1)
    ).scalar_one_or_none()

    if not latest_run:
        return {"coverage_pct": 0, "total_proposals": 0, "covered": 0, "uncovered": 0, "scopes": []}

    proposals = (
        db.execute(
            select(CenterProposal).where(
                CenterProposal.run_id == latest_run.id,
                CenterProposal.attrs["migrate"].astext == "Y",
            )
        )
        .scalars()
        .all()
    )
    total_migrating = len(proposals)

    # Get review scopes and their items
    from app.models.core import ReviewItem, ReviewScope

    scopes = db.execute(select(ReviewScope).where(ReviewScope.wave_id == wave_id)).scalars().all()

    covered_proposal_ids: set[int] = set()
    scope_details = []
    for s in scopes:
        items = db.execute(select(ReviewItem).where(ReviewItem.scope_id == s.id)).scalars().all()
        proposal_ids_in_scope = {it.proposal_id for it in items}
        covered_proposal_ids.update(proposal_ids_in_scope)
        scope_details.append(
            {
                "scope_id": s.id,
                "scope_name": s.name,
                "reviewer_name": s.reviewer_name,
                "reviewer_email": s.reviewer_email,
                "total_items": len(items),
                "has_reviewer": bool(s.reviewer_email),
            }
        )

    migrating_ids = {p.id for p in proposals}
    covered = len(covered_proposal_ids & migrating_ids)
    uncovered = total_migrating - covered

    return {
        "coverage_pct": round(covered / total_migrating * 100, 1) if total_migrating else 100,
        "total_proposals": total_migrating,
        "covered": covered,
        "uncovered": uncovered,
        "scopes": scope_details,
        "run_id": latest_run.id,
        "run_mode": latest_run.mode,
    }


# ── Unified analyse endpoint with engine selector ────────────────────────


class UnifiedAnalysisParams(BaseModel):
    """Variant-aware analysis parameters.

    The decision tree exists in two intentional variants (engines):
    - 'v1': legacy cleansing pipeline with 1:1 PC mapping
    - 'v2': CEMA migration with canonical m:1 PC grouping

    Both are configurable via ``config_id``. Inline overrides are accepted
    for V2 (pc_approach_rules / pc_start / cc_start). For V1, override the
    config via the analysis_config CRUD endpoints (fork/amend) before
    calling this endpoint.
    """

    engine: str = "v1"  # "v1" | "v2"
    config_id: int | None = None
    mode: str = "simulation"  # simulation | activated
    label: str | None = None
    excluded_scopes: list[int] | None = None
    # V2-only overrides (ignored for V1)
    pc_approach_rules: list[dict] | None = None
    pc_start: int | None = None
    cc_start: int | None = None


@router.post("/{wave_id}/analyse-with-engine")
def run_analyse_with_engine(
    wave_id: int,
    params: UnifiedAnalysisParams,
    db: Session = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "data_manager")),
) -> dict:
    """Run the configured decision-tree variant on a wave.

    This is the canonical entry point — the existing /analyse and
    /analyse-v2 endpoints stay for backward compatibility but new clients
    should use this one. Picking the engine is a deliberate choice the
    user makes per run.
    """
    wave = db.get(Wave, wave_id)
    if not wave:
        raise HTTPException(status_code=404, detail="Wave not found")
    if wave.status not in ("draft", "analysing", "proposed"):
        raise HTTPException(
            status_code=409,
            detail=f"Cannot analyse wave in status {wave.status}",
        )

    engine = (params.engine or "v1").lower()

    if engine == "v1":
        from app.services.analysis import execute_analysis, get_or_create_default_config

        config_id = params.config_id
        if config_id is None:
            config = get_or_create_default_config(db)
            config_id = config.id
        excl = [str(x) for x in (params.excluded_scopes or [])] or None
        try:
            run = execute_analysis(
                wave_id,
                config_id,
                user.id,
                db,
                mode=params.mode,
                label=params.label,
                excluded_scopes=excl,
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"V1 analysis failed: {e}") from None
        return {
            "engine": "v1",
            "run_id": run.id,
            "status": run.status,
            "kpis": run.kpis,
            "engine_version": run.engine_version,
        }

    if engine == "v2":
        import copy

        from app.models.core import AnalysisConfig
        from app.services.analysis_v2 import V2_DEFAULT_CONFIG, run_v2_analysis

        config_id = params.config_id
        # If inline V2 overrides provided, fork a runtime config
        if params.pc_approach_rules is not None or params.pc_start or params.cc_start:
            cfg = copy.deepcopy(V2_DEFAULT_CONFIG)
            if params.pc_approach_rules is not None:
                for step in cfg["pipeline"]:
                    if step["routine"] == "v2.pc_approach":
                        step["params"]["approach_rules"] = params.pc_approach_rules
            if params.pc_start is not None:
                cfg["id_assignment"]["pc_start"] = params.pc_start
            if params.cc_start is not None:
                cfg["id_assignment"]["cc_start"] = params.cc_start
            max_ver = (
                db.execute(
                    select(func.coalesce(func.max(AnalysisConfig.version), 0)).where(
                        AnalysisConfig.code == "cema_migration_v2"
                    )
                ).scalar()
                or 0
            )
            ac = AnalysisConfig(
                code="cema_migration_v2",
                version=max_ver + 1,
                name="V2 CEMA Migration (runtime)",
                config=cfg,
                created_by=user.id,
            )
            db.add(ac)
            db.flush()
            config_id = ac.id

        excl = [str(x) for x in (params.excluded_scopes or [])] or None
        try:
            result = run_v2_analysis(
                wave_id,
                config_id,
                db,
                user_id=user.id,
                mode=params.mode,
                label=params.label,
                excluded_scopes=excl,
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"V2 analysis failed: {e}") from None
        return {"engine": "v2", **result}

    raise HTTPException(
        status_code=400,
        detail=f"Unknown engine '{params.engine}'. Use 'v1' or 'v2'.",
    )
